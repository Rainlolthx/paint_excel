# coding:utf-8
# author:rain
# date  :20210609
import argparse
import os
import sys
import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from progress.bar import PixelBar
import threading
import openpyxl


def rgb2hex(r, g, b):
    return "{:02x}{:02x}{:02x}".format(r, g, b)


def paint_spreadsheet(y_max, x_max, image, filename):

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ba la la la~"

    with PixelBar("i am working on it", max=x_max) as bar:
        for col in range(1, x_max):
            for row in range(1, y_max):
                cell = ws1.cell(row, col)
                colors = image[row - 1][col - 1]
                cell.fill = PatternFill(start_color=rgb2hex(colors[2], colors[1], colors[0]), fill_type="solid")
            bar.next()
    wb.save(filename=f"{filename}.xlsx")
    return "{}.xlsx".format(filename)


def change_width_height(filelist, width=2.1, height=15.75):

    for excel in filelist:
        wb = openpyxl.load_workbook(excel)
        ws = wb[wb.sheetnames[0]]

        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = height
            
        wb.save(excel)


if __name__ == "__main__":

    parser = argparse.ArgumentParser()
    parser.add_argument("--image", help="Path to image file", required=True)
    parser.add_argument(
        "--tiny",
        action="store_true",
        help="Scale down image to 10 percent.Usefull while dealing with very high resolution images",
    )

    image_path = parser.parse_args().image
    tiny = parser.parse_args().tiny

    image = cv2.imread(image_path)
    img_w, img_h = image.shape[1], image.shape[0]
    max_h = 400
    max_w = int(max_h * (float(img_w) / img_h))
    
    max_h = max_h if max_h < img_h else img_h
    max_w = max_w if max_w < img_w else img_w
    print(img_w, img_h, max_w, max_h)

    if tiny:
        image = cv2.resize(image, (0, 0,), fx=0.1, fy=0.1)
    else:
        image = cv2.resize(image, (max_w, max_h))

    filename = paint_spreadsheet(len(image), len(image[0]), image, os.path.basename(image_path).split(".")[0])
    filelist = [filename]
    change_width_height(filelist)
    